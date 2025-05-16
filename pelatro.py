import pandas as pd
from collections import defaultdict
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

host_service_count = defaultdict(lambda: defaultdict(int))


with open('Bisous.csv', mode='r') as file:
    csv_reader = csv.DictReader(file, delimiter=';')
    for row in csv_reader:
        host = row['Host']
        service = row['Service']
        host_service_count[host][service] += 1


data = []
for host, services in host_service_count.items():
    for service, count in services.items():
        if count > 1:
            data.append([host, service, count])


df = pd.DataFrame(data, columns=['Host', 'Service', 'Count'])


wb = Workbook()
ws = wb.active
ws.title = "Service Count"


headers = ['Host', 'Service', 'Count']
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment


for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)


for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width


wb.save('REDONDANT ALERT PELATRO.xlsx')

print("List créé avec succès et enregistré sous le nom 'REDONDANT ALERT PELATRO'.")